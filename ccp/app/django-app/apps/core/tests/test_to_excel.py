"""Tests for :mod:`apps.core.exports.to_excel`."""

from __future__ import annotations

import io

import openpyxl
import pytest

from apps.core.exports.to_excel import (
    SHEET_GARANTIA,
    SHEET_GRAFICOS,
    SHEET_PARAMETROS,
    SHEET_RESULTADOS,
    SHEET_TESTE,
    session_to_excel,
)


@pytest.fixture()
def sample_state() -> dict:
    return {
        "parameters": {
            "speed": {"value": 12000, "units": "rpm"},
            "rotor_name": {"value": "R1", "units": None},
        },
        "guarantee_point": {
            "flow_m": 15.0,
            "ps": 1.0,
            "Ts": 300.0,
            "pd": 3.0,
            "Td": 420.0,
        },
        "test_points": [
            {"Nome": "T1", "flow_m": 14.0, "ps": 1.0, "Ts": 300.0},
            {"Nome": "T2", "flow_m": 16.0, "ps": 1.0, "Ts": 300.0},
        ],
        "results": [
            {"Ponto": "T1", "Eficiencia": 0.82, "Head": 50.0},
            {"Ponto": "T2", "Eficiencia": 0.80, "Head": 48.5},
        ],
        "plots": {"head": b"PNGFAKE", "eff": b"PNGFAKE2"},
    }


def test_session_to_excel_returns_bytes(sample_state):
    payload = session_to_excel(sample_state)
    assert isinstance(payload, bytes)
    assert payload.startswith(b"PK")  # ZIP magic — xlsx is a zip container.


def test_session_to_excel_sheets_present(sample_state):
    payload = session_to_excel(sample_state)
    wb = openpyxl.load_workbook(io.BytesIO(payload))
    names = set(wb.sheetnames)
    assert {
        SHEET_PARAMETROS,
        SHEET_GARANTIA,
        SHEET_TESTE,
        SHEET_RESULTADOS,
        SHEET_GRAFICOS,
    } <= names


def test_session_to_excel_preserves_portuguese_columns(sample_state):
    payload = session_to_excel(sample_state)
    wb = openpyxl.load_workbook(io.BytesIO(payload))
    params = wb[SHEET_PARAMETROS]
    header_row = [cell.value for cell in next(params.iter_rows(min_row=1, max_row=1))]
    assert "Parametro" in header_row
    assert "Valor" in header_row
    assert "Unidade" in header_row

    plots = wb[SHEET_GRAFICOS]
    plot_header = [cell.value for cell in next(plots.iter_rows(min_row=1, max_row=1))]
    assert "Grafico" in plot_header


def test_session_to_excel_empty_state():
    payload = session_to_excel({})
    wb = openpyxl.load_workbook(io.BytesIO(payload))
    assert SHEET_PARAMETROS in wb.sheetnames
    assert SHEET_RESULTADOS in wb.sheetnames
