"""Module to define impeller class."""

import csv
import multiprocessing
import warnings

import toml
from copy import deepcopy
from itertools import groupby
from pathlib import Path

import numpy as np
import plotly.graph_objects as go
from openpyxl import Workbook
from scipy.interpolate import interp1d, UnivariateSpline, PchipInterpolator
from scipy.optimize import fsolve

from ccp import Q_, State, Point, Curve
from ccp.config.units import check_units
from ccp.config.utilities import r_getattr, r_setattr
from ccp.data_io.read_csv import read_data_from_engauge_csv
from ccp.plotly_theme import tableau_colors


class ImpellerStateParameter:
    def __init__(self, impeller_state_object, attr):
        self.impeller_state_object = impeller_state_object
        self.attr = attr

    def __call__(self, *args, **kwargs):
        impeller_state_object = self.impeller_state_object
        attr = self.attr
        values = []

        for curve_state in impeller_state_object:
            values.append(getattr(getattr(curve_state, attr)(), "magnitude"))

        units = getattr(getattr(curve_state, attr)(), "units")

        return Q_(values, units)


def impeller_state_parameter(impeller_state_object, attr):
    return ImpellerStateParameter(impeller_state_object, attr)


class ImpellerState:
    def __init__(self, curves_state):
        self.curves_state = curves_state

        for attr in ["p", "T", "h", "s", "rho"]:
            func = impeller_state_parameter(self, attr)
            setattr(self, attr, func)

    def __getitem__(self, item):
        return self.curves_state.__getitem__(item)


class ImpellerPlotFunction:
    def __init__(self, impeller_object, attr):
        self.impeller_object = impeller_object
        self.attr = attr

    @check_units
    def __call__(
        self,
        *args,
        flow_v=None,
        speed=None,
        speed_units="RPM",
        plot_kws=None,
        show_points=False,
        **kwargs,
    ):
        """Plot parameter versus volumetric flow.

        You can plot a specific point in the plot giving its flow_v and speed.

        You can choose units with the arguments flow_v_units='...' and
        {attr}_units='...'. For the speed you can use speed_units='...'.

        Parameters
        ----------
        flow_v : pint.Quantity, float, optional
            Volumetric flow (m³/s) for a specific point in the plot.
        speed : pint.Quantity, float, optional
            Speed (rad/s) for a specific point in the plot.
        flow_v_units : str, optional
            Flow units used for the plot. Default is m³/s.
        {attr}_units : str, optional
            Units for the parameter being plotted (e.g. for a head plot we could use
            head_units='J/kg' or head_units='J/g'). Default is SI.
        speed_units : str, optional
            Speed units for the plot. Default is 'RPM'.
        show_points : bool, optional
            If True, the points will be plotted as markers.

        Returns
        -------
        fig : plotly.Figure
            Plotly figure that can be customized.

        Examples
        --------
        >>> import ccp
        >>> imp = ccp.impeller_example()
        >>> fig = imp.head_plot(
        ...    flow_v=5.5,
        ...    speed=900,
        ...    flow_v_units='m³/h',
        ...    head_units='J/kg',
        ...    speed_units='RPM'
        ... )

        """
        impeller_object = self.impeller_object
        attr = self.attr
        fig = kwargs.pop("fig", None)

        if fig is None:
            fig = go.Figure()

        if plot_kws is None:
            plot_kws = {}

        p0 = impeller_object.points[0]
        flow_v_units = kwargs.get("flow_v_units", p0.flow_v.units)

        # store values for surge flow and attr value in lists so that we can plot them
        surge_flow_list = []
        surge_attr_list = []

        if "." in attr:
            attr_str = attr.split(".")[-1]
        else:
            attr_str = attr
        
        try:
            attr_units = kwargs.get(
                f"{attr_str}_units", r_getattr(impeller_object.curves[0], attr).units
            )
        except AttributeError:
            attr_units = kwargs.get(
                f"{attr_str}_units", r_getattr(impeller_object.curves[0], attr)().units
            )

        color_iterator = iter(tableau_colors)
        for curve in impeller_object.curves:
            color = next(color_iterator)
            fig = r_getattr(curve, attr + "_plot")(
                fig=fig,
                speed_units=speed_units,
                plot_kws=plot_kws,
                color=color,
                show_points=show_points,
                **kwargs,
            )
            surge_flow_list.append(curve.flow_v[0].to(flow_v_units).m)
            try:
                surge_attr_list.append(r_getattr(curve, attr)[0].to(attr_units).m)
            except (AttributeError, TypeError):
                surge_attr_list.append(r_getattr(curve, attr)()[0].to(attr_units).m)

        if speed:
            color = next(color_iterator)
            current_curve = impeller_object.curve(speed=speed)
            fig = r_getattr(current_curve, attr + "_plot")(
                fig=fig,
                plot_kws=plot_kws,
                color=color,
                speed_units=speed_units,
                show_points=show_points,
                **kwargs,
            )

            color = "black"
            if flow_v:
                current_point = impeller_object.point(flow_v=flow_v, speed=speed)
                fig = r_getattr(current_point, attr + "_plot")(
                    fig=fig,
                    speed_units=speed_units,
                    plot_kws=plot_kws,
                    color=color,
                    **kwargs,
                )

        # add surge flow and attr values to plot as dotted line
        fig.add_trace(
            go.Scatter(
                x=surge_flow_list,
                y=surge_attr_list,
                mode="lines",
                # gray line with transparency
                line=dict(color="black", dash="dash"),
                showlegend=False,
                opacity=0.3,
            )
        )

        # extra x range
        flow_values = [p.flow_v.to(flow_v_units) for p in impeller_object.points]
        min_flow = min(flow_values)
        max_flow = max(flow_values)
        delta = 0.05 * (max_flow - min_flow)
        fig.update_layout(xaxis=dict(range=[min_flow - delta, max_flow + delta]))

        return fig


def impeller_plot_function(impeller_object, attr):
    return ImpellerPlotFunction(impeller_object, attr)


class CompareImpellerPlotFunction:
    def __init__(self, impeller_object, attr):
        self.impeller_object = impeller_object
        self.attr = attr

    @check_units
    def __call__(
        self,
        other_impeller,
        flow_v=None,
        speed=None,
        speed_units="RPM",
        plot_kws=None,
        **kwargs,
    ):
        """Plot parameter versus volumetric flow.

        You can plot a specific point in the plot giving its flow_v and speed.

        You can choose units with the arguments flow_v_units='...' and
        {attr}_units='...'. For the speed you can use speed_units='...'.

        Parameters
        ----------
        flow_v : pint.Quantity, float, optional
            Volumetric flow (m³/s) for a specific point in the plot.
        speed : pint.Quantity, float, optional
            Speed (rad/s) for a specific point in the plot.
        flow_v_units : str, optional
            Flow units used for the plot. Default is m³/s.
        {attr}_units : str, optional
            Units for the parameter being plotted (e.g. for a head plot we could use
            head_units='J/kg' or head_units='J/g'). Default is SI.
        speed_units : str, optional
            Speed units for the plot. Default is 'RPM'.

        Returns
        -------
        fig : plotly.Figure
            Plotly figure that can be customized.

        Examples
        --------
        >>> import ccp
        >>> imp = ccp.impeller_example()
        >>> fig = imp.plot_head(
        ...    flow_v=5.5,
        ...    speed=900,
        ...    flow_v_units='m³/h',
        ...    head_units='j/kg',
        ...    speed_units='RPM'
        ... )

        """
        impeller_object = self.impeller_object
        attr = self.attr
        fig = kwargs.pop("fig", None)

        if fig is None:
            fig = go.Figure()

        if plot_kws is None:
            plot_kws = {}

        p0 = impeller_object.points[0]
        flow_v_units = kwargs.get("flow_v_units", p0.flow_v.units)

        for curve, color in zip(impeller_object.curves, tableau_colors):
            other_curve = other_impeller.curve(speed=curve.speed)
            fig = r_getattr(other_curve, attr + "_plot")(
                fig=fig, speed_units=speed_units, plot_kws=plot_kws, **kwargs
            )
            # change line to dash
            fig.data[-1].update(
                line=dict(dash="dash", color=color),
                showlegend=False,
            )

        for curve, color in zip(impeller_object.curves, tableau_colors):
            fig = r_getattr(curve, attr + "_plot")(
                fig=fig, speed_units=speed_units, plot_kws=plot_kws, **kwargs
            )
            fig.data[-1].update(line=dict(color=color))

        if speed:
            current_curve = impeller_object.curve(speed=speed)
            fig = r_getattr(current_curve, attr + "_plot")(
                fig=fig, speed_units=speed_units, plot_kws=plot_kws, **kwargs
            )
            if flow_v:
                current_point = impeller_object.point(flow_v=flow_v, speed=speed)
                fig = r_getattr(current_point, attr + "_plot")(
                    fig=fig, speed_units=speed_units, plot_kws=plot_kws, **kwargs
                )

        # extra x range
        flow_values = [p.flow_v.to(flow_v_units) for p in impeller_object.points]
        min_flow = min(flow_values)
        max_flow = max(flow_values)
        delta = 0.05 * (max_flow - min_flow)
        fig.update_layout(xaxis=dict(range=[min_flow - delta, max_flow + delta]))

        return fig


def compare_impeller_plot_function(impeller_object, attr):
    return CompareImpellerPlotFunction(impeller_object, attr)


class Impeller:
    """An impeller with a performance map.

    Impeller instance is initialized with the list of points.
    The created instance will hold the dimensional points used in instantiation.
    Curves will be generated from points close in similarity.

    Parameters
    ----------
    points : list
        List with ccp.Point objects.

    Returns
    -------
    impeller : ccp.Impeller
    """

    @check_units
    def __init__(self, points):
        self.points = deepcopy(points)

        losses_dict = {p.power_losses: p.speed for p in self.points}
        max_losses = max(losses_dict.keys())
        max_losses_speed = losses_dict[max_losses]

        curves = []
        for speed, grouped_points in groupby(
            sorted(self.points, key=lambda point: point.speed),
            key=lambda point: point.speed,
        ):
            if max_losses.m > 0:
                points = []
                for p in grouped_points:
                    if p.power_losses.m == 0:
                        losses = max_losses * (p.speed / max_losses_speed) ** 2.5
                        p_new = Point(
                            suc=p.suc,
                            disch=p.disch,
                            flow_v=p.flow_v,
                            speed=p.speed,
                            power_losses=losses,
                            b=p.b,
                            D=p.D,
                        )
                    else:
                        p_new = deepcopy(p)
                    points.append(p_new)
            else:
                points = [point for point in grouped_points]
            curve = Curve(points)
            curves.append(curve)
            setattr(self, f"curve_{int(curve.speed.magnitude)}", curve)
        self.curves = curves
        self.disch = ImpellerState([c.disch for c in self.curves])

        for attr in [
            "disch.p",
            "disch.T",
            "disch.h",
            "disch.s",
            "disch.rho",
            "head",
            "eff",
            "power",
            "power_shaft",
            "torque",
            "psi",
            "phi",
            "flow_v",
            "flow_m",
            "speed",
        ]:
            values = []
            # for disch.p etc values are defined in _Impeller_State
            if "." not in attr:
                for c in self.curves:
                    param = r_getattr(c, attr)
                    values.append(param.magnitude)
                units = param.units
                r_setattr(self, attr, Q_(values, units))

            r_setattr(self, f"{attr}_plot", impeller_plot_function(self, attr))
            r_setattr(
                self, f"{attr}_compare", compare_impeller_plot_function(self, attr)
            )

    def __getitem__(self, item):
        return self.points.__getitem__(item)

    def __eq__(self, other):
        if isinstance(other, self.__class__):
            points_other = sorted(other.points, key=lambda x: x.flow_v)
            points_self = sorted(self.points, key=lambda x: x.flow_v)
            if len(points_other) != len(points_self):
                return False
            else:
                return points_other == points_self

    @check_units
    def point(self, flow_v=None, flow_m=None, speed=None):
        """Calculate specific point in the performance map.

        Given a volumetric flow and a speed this method will calculate a point in the
        impeller map according to these arguments.

        Parameters
        ----------
        flow_v : pint.Quantity, float
            Volumetric flow (m³/s).
        flow_m : pint.Quantity, float
            Mass flow (kg/s).
        speed : pint.Quantity, float
            Speed (rad/s).

        Returns
        -------
        point : ccp.Point
            Point in the performance map.
        """

        # check if speed and flow are defined
        if speed is None:
            raise ValueError("Speed must be defined.")
        if flow_v is None and flow_m is None:
            raise ValueError("Either flow_v or flow_m must be defined.")

        current_curve = self.curve(speed)
        if flow_m:
            flow_v = current_curve.points[0].suc.v() * flow_m

        func_T = interp1d(
            current_curve.flow_v.m, current_curve.disch.T().m, fill_value="extrapolate"
        )
        func_p = interp1d(
            current_curve.flow_v.m, current_curve.disch.p().m, fill_value="extrapolate"
        )

        min_flow_v = min(current_curve.flow_v)
        max_flow_v = max(current_curve.flow_v)
        if flow_v < min_flow_v or max_flow_v < flow_v:
            warnings.warn(
                f"Expected point is being extrapolated.\n"
                f"Interpolation limits: {min_flow_v:.3f~P} ~ {max_flow_v:.3f~P}\n"
                f"Expected point flow: {flow_v:.3f~P}"
            )

        flow_at_min_p = (
            np.log(current_curve[-1].disch.p().m + np.exp(4 * max_flow_v.m))
        ) / 4
        flow_at_min_T = (
            np.log(
                current_curve[-1].disch.T().m
                - current_curve.points[0].suc.T().m
                + np.exp(4 * max_flow_v.m)
            )
        ) / 4

        # Extrapolation code for choke region
        if flow_v <= max_flow_v:
            disch_p = func_p(flow_v)
        elif flow_v.m < flow_at_min_p:
            disch_p = round(
                current_curve[-1].disch.p().m
                + np.exp(4 * current_curve[-1].flow_v.m)
                - np.exp(4 * flow_v.m),
                2,
            )
        else:
            disch_p = 0.001

        if flow_v <= max_flow_v:
            disch_T = func_T(flow_v)
        elif flow_v.m < flow_at_min_T:
            disch_T = round(
                current_curve[-1].disch.T().m
                + np.exp(4 * current_curve[-1].flow_v.m)
                - np.exp(4 * flow_v.m),
                2,
            )
        else:
            disch_T = current_curve.points[0].suc.T().m

        p0 = self.points[0]
        disch = State(p=disch_p, T=disch_T, fluid=p0.suc.fluid)

        power_losses = current_curve.power_losses

        point = Point(
            suc=p0.suc,
            disch=disch,
            flow_v=flow_v,
            speed=current_curve.speed,
            b=p0.b,
            D=p0.D,
            power_losses=power_losses,
        )

        return point

    @check_units
    def curve(self, speed=None):
        """Calculate specific point in the performance map.

        Given a speed this method will calculate a curve in the
        impeller map according to these arguments.

        Parameters
        ----------
        speed : pint.Quantity, float
            Speed (rad/s).

        Returns
        -------
        curve : ccp.Curve
            Point in the performance map.
        """
        speeds = np.array([curve.speed.magnitude for curve in self.curves])

        # handle case where we only have one curve and can't interpolate
        if len(speeds) == 1:
            current_curve = self.curves[0]
            if speed is not None and not np.allclose(speed, current_curve.speed):
                raise ValueError(
                    f"Can only interpolate for speed={current_curve.speed}"
                )
            return current_curve

        # calculate power losses
        power_losses = calculate_power_losses(
            power_losses_ref=self.curves[0].power_losses,
            speed_ref=self.curves[0].speed,
            speed=speed,
        )

        closest_curves_idxs = find_closest_speeds(speeds, speed.magnitude)
        curves = [
            self.curves[closest_curves_idxs[0]],
            self.curves[closest_curves_idxs[1]],
        ]

        # calculate factor
        speed_range = curves[1].speed.magnitude - curves[0].speed.magnitude
        factor = (speed.magnitude - curves[0].speed.magnitude) / speed_range

        current_curve = []
        p0 = self.points[0]
        number_of_points = len(curves[0])

        for i in range(number_of_points):
            flow_T, disch_T = get_interpolated_values(
                factor,
                curves[0].flow_v.magnitude[i],
                curves[0][i].disch.T().m,
                curves[1].flow_v.magnitude[i],
                curves[1][i].disch.T().m,
            )
            flow_p, disch_p = get_interpolated_values(
                factor,
                curves[0].flow_v.magnitude[i],
                curves[0][i].disch.p().m,
                curves[1].flow_v.magnitude[i],
                curves[1][i].disch.p().m,
            )

            disch = State(p=disch_p, T=disch_T, fluid=p0.suc.fluid)

            p = Point(
                suc=p0.suc,
                disch=disch,
                flow_v=(flow_T + flow_p) / 2,
                speed=speed,
                power_losses=power_losses,
                b=p0.b,
                D=p0.D,
            )

            current_curve.append(p)

        current_curve = Curve(current_curve)

        return current_curve

    @classmethod
    def convert_from(cls, original_impeller, suc=None, find="speed", speed=None):
        """Convert performance map from an impeller.

        Parameters
        ----------
        original_impeller : ccp.Impeller, list
            The original impeller.
            A list of impeller can also be passed. In this case the curves will be
            converted based on the impeller with the closest suction speed of sound
            to the new suction condition.
        suc : ccp.State
            The new suction condition to which we want to convert to.
        find : str, optional
            The method in which the curves will be converted.
            For now only 'speed' is implemented, which means that, based on volume ratio,
            we calculate new values of speed for each curve and the respective discharge
            condition.
        speed : float, pint.Quantity, str, optional
            Desired speed. If find="speed", this can be None or 'same' to keep the same
            speed values available in the original_impeller.

        Returns
        -------
        converted_impeller : ccp.Impeller
            The new impeller with the converted performance map for the required
            suction condition.
        """
        all_converted_points = []
        if isinstance(original_impeller, list):
            speed_sound_diff = []
            for impeller in original_impeller:
                speed_sound_diff.append(
                    impeller.points[0].suc.speed_sound().m - suc.speed_sound().m
                )
            original_impeller = original_impeller[np.argmin(np.abs(speed_sound_diff))]

        for curve in original_impeller.curves:
            with multiprocessing.Pool() as pool:
                converter_args = [(p, suc, find) for p in curve]
                converted_points = pool.map(converter, converter_args)

                if speed is None or speed == "same":
                    speed_mean = np.mean([p.speed.magnitude for p in converted_points])
                else:
                    speed_mean = speed

                converted_points = [
                    Point.convert_from(
                        p,
                        suc=p.suc,
                        find="volume_ratio",
                        speed=speed_mean,
                    )
                    for p in converted_points
                ]

                all_converted_points += converted_points

        converted_impeller = cls(all_converted_points)
        if speed == "same":
            all_converted_points = []
            for curve in original_impeller.curves:
                converted_curve = converted_impeller.curve(curve.speed)
                all_converted_points += converted_curve.points

            converted_impeller = cls(all_converted_points)

        return converted_impeller

    def _calc_new_points(self):
        """Calculate new dimensional points based on the suction condition."""
        #  keep volume ratio constant
        all_points = []
        for curve in self.curves:
            new_points = []
            for point in curve:
                new_points.append(self._calc_from_non_dimensional(point))

            speed_mean = np.mean([p.speed.magnitude for p in new_points])

            new_points = [
                self._calc_from_speed(p, Q_(speed_mean, p.speed.units))
                for p in new_points
            ]

            all_points += new_points

        self.new = self.__class__(
            all_points,
            b=self.b,
            D=self.D,
            _suc=self.new_suc,
            _flow_v=self.flow_v,
            _speed=self.speed,
        )

    def export_to_excel(self, path_name=None):
        """Export curves to excel file."""
        wb = Workbook()
        for curve in self.curves:
            sheet_name = f'{curve.speed.to("RPM"):.0f~P}'
            ws = wb.create_sheet(sheet_name)
            for i, p in enumerate(curve):
                i += 1  # openpyxl index
                if i == 1:
                    ws.cell(row=i, column=1, value="Flow (m**3/s)")
                    ws.cell(row=i, column=2, value="Head (kJ/kg)")
                    ws.cell(row=i, column=3, value="Efficiency (%)")
                else:
                    ws.cell(row=i, column=1, value=p.flow_v.magnitude)
                    ws.cell(row=i, column=2, value=p.head.to("kJ/kg").magnitude)
                    ws.cell(row=i, column=3, value=p.eff.magnitude * 100)

        if path_name is None:
            file_name = f'{self.suc.p().to("bar"):.0f~P}.xlsx'
            file_name = file_name.replace(" ", "-")
            path_name = Path.cwd() / file_name

        wb.save(str(path_name))

    @classmethod
    def load_from_dict(
        cls,
        suc,
        head_curves=None,
        eff_curves=None,
        power_curves=None,
        power_shaft_curves=None,
        pressure_ratio_curves=None,
        disch_T_curves=None,
        b=Q_(0.005, "m"),
        D=Q_(0.5, "m"),
        number_of_points=10,
        flow_units="m**3/s",
        flow_units_head=None,
        flow_units_eff=None,
        flow_units_power=None,
        flow_units_power_shaft=None,
        flow_units_pressure_ratio=None,
        flow_units_disch_T=None,
        head_units="J/kg",
        eff_units="dimensionless",
        pressure_ratio_units="dimensionless",
        disch_T_units="degK",
        power_units="W",
        power_shaft_units="W",
        power_losses_units="W",
        speed_units="RPM",
    ):
        """Create points from dict object.

        In this case the dict is in the following format:

        curve = {
            '10000': {
                'x1': [list with flow values],
                'x2': [list with head or eff values]
                'x3': power losses value},
            ...}
        Where 10000 is the speed for that curve.

        suc : ccp.State
            Suction state.
        head_curves : dict
            Dict with head/flow values.
        eff_curves : dict
            Dict with eff/flow values.
        power_curves : dict
            Dict with power/flow values.
        b : float, pint.Quantity
            Impeller width at the outer blade diameter (m).
        D : float, pint.Quantity
            Impeller outer diameter (m).
        number_of_points : int
            Number of points that will be interpolated.
        flow_units : str
            Flow units used in the dict.
        flow_units_head: str
            Flow units used in the dict for head curves.
            Only needed when flow units for head curve differs from other curves.
        flow_units_eff: str
            Flow units used  in the dict for efficiency curves.
            Only needed when flow units for efficiency curve differs from other curves.
        flow_units_power: str
            Flow units used  in the dict for power curves.
            Only needed when flow units for power curve differs from other curves.
        flow_units_power_shaft: str
            Flow units used in the dict for shaft power curves.
            Only needed when flow units for power curve differs from other curves.
        flow_units_pressure_ratio: str
            Flow units used  in the dict for pressure ratio curves.
            Only needed when flow units for efficiency curve differs from other curves.
        flow_units_disch_T: str
            Flow units used  in the dict for discharge Temperature curves.
            Only needed when flow units for efficiency curve differs from other curves.
        head_units : str
            Head units used in the dict.
            If the curve head units are in meter you can use: head_units="m*g0".
        eff_units : str
            Dimensionless.
        power_units : str
            Power units used in the dict.
        power_shaft_units : str
            Shaft power units used in the dict.
        power_losses_units : str
            Mechanical power losses units used in the dict.
        pressure_ratio_units : str
            Pressure ratio units used in the dict.
        disch_T_units : str
            Discharge temperature units used in the dict.
        speed_units : str
            Speed units used in the dict.
        """
        # define if we have volume or mass flow
        args = locals().copy()

        flow_type = "volumetric"
        if list(Q_(1, flow_units).dimensionality.keys())[0] == "[mass]":
            flow_type = "mass"

        points = []

        curves = {}
        for k, v in args.items():
            if "curves" in k and v is not None:
                name_split = k.split("_")
                try:
                    curve_name = "_".join(name_split[:-1])
                except TypeError:
                    curve_name = name_split[0]
                curves[curve_name] = v

        parameters = list(curves)
        speeds = list(curves[parameters[0]])

        for param in parameters:
            if not args[f"flow_units_{param}"]:
                args[f"flow_units_{param}"] = flow_units

        flow_units_list = {
            f"flow_units_{param}": (
                "mass"
                if list(Q_(1, args[f"flow_units_{param}"]).dimensionality.keys())[0]
                == "[mass]"
                else "volumetric"
            )
            for param in parameters
        }

        if "eff" in curves:
            if max(curves["eff"][speeds[0]]["x2"]) > 1:
                for speed in speeds:
                    curves["eff"][speed]["x2"] = [
                        i / 100 for i in curves["eff"][speed]["x2"]
                    ]

        # create interpolated curves
        curves_interpolated = {}
        # dict to hold interpolated values for specific x values
        points_interpolated = {}

        for speed in speeds:
            min_x = 0
            max_x = 1e20
            # get min and max flow
            for n, curve in enumerate(curves):
                if n == 0:
                    power_losses = curves[curve][speed]["x3"]
                else:
                    if power_losses != curves[curve][speed]["x3"]:
                        raise ValueError(
                            f"There should be the same power losses values, for the same speed. \n"
                            f"Currently we have different values for {speed.to(speed_units):.2f~P} "
                            f"{speed_units}. \n"
                            "Please check and try again."
                        )
                if flow_units_list[f"flow_units_{curve}"] != flow_type:
                    if flow_units_list[f"flow_units_{curve}"] == "mass":
                        flow_curve = [
                            Q_(flow, args[f"flow_units_{curve}"]) / suc.rho()
                            for flow in curves[curve][speed]["x1"]
                        ]
                        curves[curve][speed]["x1"] = [
                            flow.to(flow_units).m for flow in flow_curve
                        ]
                    else:
                        flow_curve = [
                            Q_(flow, args[f"flow_units_{curve}"]) * suc.rho()
                            for flow in curves[curve][speed]["x1"]
                        ]
                        curves[curve][speed]["x1"] = [
                            flow.to(flow_units).m for flow in flow_curve
                        ]
                else:
                    flow_curve = [
                        Q_(flow, args[f"flow_units_{curve}"])
                        for flow in curves[curve][speed]["x1"]
                    ]
                    curves[curve][speed]["x1"] = [
                        flow.to(flow_units).m for flow in flow_curve
                    ]
                if curves[curve][speed]["x1"][0] > min_x:
                    min_x = curves[curve][speed]["x1"][0]
                if curves[curve][speed]["x1"][-1] < max_x:
                    max_x = curves[curve][speed]["x1"][-1]

            points_x = np.linspace(min_x, max_x, number_of_points)

            for curve in curves:
                curves_interpolated[curve] = {}
                points_interpolated[curve] = {}

                curves_interpolated[curve][speed] = PchipInterpolator(
                    curves[curve][speed]["x1"],
                    curves[curve][speed]["x2"],
                    extrapolate=True,
                )
                points_interpolated[curve][speed] = curves_interpolated[curve][speed](
                    points_x
                )

            args_list = []

            for flow, param0, param1 in zip(
                points_x,
                points_interpolated[parameters[0]][speed],
                points_interpolated[parameters[1]][speed],
            ):
                arg_dict = {
                    "suc": suc,
                    "speed": Q_(float(speed), speed_units),
                    parameters[0]: Q_(param0, args[f"{parameters[0]}_units"]),
                    parameters[1]: Q_(param1, args[f"{parameters[1]}_units"]),
                    "power_losses": Q_(power_losses, power_losses_units),
                    "b": b,
                    "D": D,
                }
                if flow_type == "volumetric":
                    arg_dict["flow_v"] = Q_(flow, flow_units)
                elif flow_type == "mass":
                    arg_dict["flow_m"] = Q_(flow, flow_units)
                args_list.append(arg_dict)

            with multiprocessing.Pool() as pool:
                points += pool.map(create_points_parallel, args_list)

        return cls(points)

    @classmethod
    def load_from_dict_isis(
        cls,
        suc,
        head_curves,
        eff_curves,
        b=Q_(0.005, "m"),
        D=Q_(0.5, "m"),
        number_of_points=10,
        flow_units="m**3/s",
        head_units="J/kg",
        speed_units="RPM",
    ):
        """Create points from dict object available in the ISIS platform.

        The dict is in the following format:
           head_curves_dict = [
        {
            "z": 11373,
            "points": [
                {"x": 94529, "y": 148.586},
                {"x": 98641, "y": 148.211},
                {"x": 101554, "y": 147.837},
            ...]
        ...,
        },
        ]

        Where z is the speed and each point is described as a dict with x and y
        pair where x is the flow and y is the head or eff.

        suc : ccp.State
            Suction state.
        head_curve : dict
            Dict with head/flow values.
        eff_curve : dict
            Dict with head/flow values.
        b : float, pint.Quantity
            Impeller width at the outer blade diameter (m).
        D : float, pint.Quantity
            Impeller outer diameter (m).
        number_of_points : int
            Number of points that will be interpolated.
        flow_units : str
            Flow units used in the dict.
        head_units : str
            Head units used in the dict.
            If the curve head units are in meter you can use: head_units="m*g0".
        speed_units : str
            Speed units used in the dict.

        Examples
        --------
        >>> import ccp
        >>> composition_fd = dict(
        ...    n2=0.4,
        ...    co2=0.22,
        ...    methane=92.11,
        ...    ethane=4.94,
        ...    propane=1.71,
        ...    ibutane=0.24,
        ...    butane=0.3,
        ...    ipentane=0.04,
        ...    pentane=0.03,
        ...    hexane=0.01,
        ... )
        >>> suc_fd = State(p=Q_(3876, "kPa"), T=Q_(11, "degC"), fluid=composition_fd)
        >>> imp = ccp.Impeller.load_from_dict_isis(
        ...    suc=suc_fd,
        ...    head_curves=head_curves_dict,
        ...    eff_curves=eff_curves_dict,
        ...    b=Q_(10.6, "mm"),
        ...    D=Q_(390, "mm"),
        ...    number_of_points=6,
        ...    flow_units="kg/h",
        ...    head_units="kJ/kg",
        ... )
        """
        # change dict format from isis to that handled by the ccp method.

        head_curve_ccp = {}
        for head_curve in head_curves["CURVES"]:
            speed = str(head_curve["z"])
            x = []
            y = []
            for point in head_curve["points"]:
                x.append(point["x"])
                y.append(point["y"])

            head_curve_ccp[speed] = {}
            head_curve_ccp[speed]["x1"] = x
            head_curve_ccp[speed]["x2"] = y
            head_curve_ccp[speed]["x3"] = 0

        eff_curve_ccp = {}
        for eff_curve in eff_curves["CURVES"]:
            speed = str(eff_curve["z"])
            x = []
            y = []
            for point in eff_curve["points"]:
                x.append(point["x"])
                y.append(point["y"])

            eff_curve_ccp[speed] = {}
            eff_curve_ccp[speed]["x1"] = x
            eff_curve_ccp[speed]["x2"] = y
            eff_curve_ccp[speed]["x3"] = 0

        return cls.load_from_dict(
            suc=suc,
            b=b,
            D=D,
            head_curves=head_curve_ccp,
            eff_curves=eff_curve_ccp,
            number_of_points=number_of_points,
            flow_units=flow_units,
            head_units=head_units,
            speed_units=speed_units,
        )

    @classmethod
    def load_from_engauge_csv(
        cls,
        suc,
        curve_name,
        curve_path,
        b=Q_(0.005, "m"),
        D=Q_(0.5, "m"),
        number_of_points=10,
        flow_units="m**3/s",
        flow_units_head=None,
        flow_units_eff=None,
        flow_units_power=None,
        flow_units_power_shaft=None,
        flow_units_pressure_ratio=None,
        flow_units_disch_T=None,
        head_units="J/kg",
        eff_units="dimensionless",
        power_units="W",
        power_shaft_units="W",
        power_losses_units="W",
        pressure_ratio_units="dimensionless",
        disch_T_units="degK",
        speed_units="RPM",
        **kwargs,
    ):
        """Convert points from csv generated by engauge to csv with 6 points at same flow for use on hysys.

        The csv files should be generated with engauge with the following procedure:
        First, copy the image of the curve to your clipboard, then inside engauge digitizer:
            - Edit -> Paste as new
            - Name each curve with their respective speed value (e.g. 10322);
            - If the curve is for shaft power, you can add the power loss next to the speed value (e.g. 10322, 82);
            - On Axis Point -> add 3 reference points
            - Select the curve (e.g. 10322 would be the curve for 10322 RPM)
            - Select the points using the segment fill tool;
            - Select the next curve and points...
            - Settings -> Export Setup -> Select:
            - Raws X's and Y's ; One curve on each line
        Files should be saved with the following convention:
            - <curve-name>-head.csv
            - <curve-name>-eff.csv

        suc : ccp.State
            Suction state.
        curve_path : pathlib.Path
            Path to the curves.
        curve_name : str
            Name for head and efficiency curve.
            Curves should have names <curve_name>-head.csv and <curve-name>-eff.csv.
        b : float, pint.Quantity
            Impeller width at the outer blade diameter (m).
        D : float, pint.Quantity
            Impeller outer diameter (m).
        number_of_points : int
            Number of points that will be interpolated.
        flow_units : str
            Flow units used when extracting data with engauge.
        flow_units_head: str
            Flow units used when extracting data with engauge for head curves.
            Only needed when flow units for head curve differs from other curves.
        flow_units_eff: str
            Flow units used when extracting data with engauge for efficiency curves.
            Only needed when flow units for efficiency curve differs from other curves.
        flow_units_power: str
            Flow units used when extracting data with engauge for power curves.
            Only needed when flow units for power curve differs from other curves.
        flow_units_power_shaft: str
            Flow units used when extracting data with engauge for shaft power curves.
            Only needed when flow units for power curve differs from other curves.
        flow_units_pressure_ratio: str
            Flow units used when extracting data with engauge for pressure ratio curves.
            Only needed when flow units for power curve differs from other curves.
        flow_units_disch_T: str
            Flow units used when extracting data with engauge for discharge temperature curves.
            Only needed when flow units for efficiency curve differs from other curves.
        head_units : str
            Head units used when extracting data with engauge.
            If the curve head units are in meter you can use: head_units="m*g0".
        eff_units : str
            Dimensionless.
        power_units : str
            Power units used when extracting data with engauge.
        power_shaft_units : str
            Shaft power units used when extracting data with engauge.
        power_losses_units : str
            Power losses units used when extracting data with engauge.
        pressure_ratio_units : str
            Pressure ratio units used when extracting data with engauge.
        disch_T_units : str
            Discharge temperature units used when extracting data with engauge.
        speed_units : str
            Speed units used when extracting data with engauge.
        """
        curves_path_dict = {}

        for param in [
            "head",
            "eff",
            "power",
            "power_shaft",
            "pressure_ratio",
            "disch_T",
        ]:
            param_path = curve_path / (curve_name + f"-{param}.csv")
            if param_path.is_file():
                curves_path_dict[f"{param}_curves"] = read_data_from_engauge_csv(
                    param_path
                )

        if len(curves_path_dict) != 2:
            raise ValueError(
                f"There should be 2 curves, currently we have: \n"
                f" {[curve.removesuffix('_curves') for curve in curves_path_dict.keys()]}\n"
                "Please check the file names and path and try again."
            )

        return cls.load_from_dict(
            suc=suc,
            b=b,
            D=D,
            number_of_points=number_of_points,
            flow_units=flow_units,
            head_units=head_units,
            power_units=power_units,
            power_shaft_units=power_shaft_units,
            power_losses_units=power_losses_units,
            flow_units_head=flow_units_head,
            flow_units_eff=flow_units_eff,
            flow_units_power=flow_units_power,
            flow_units_power_shaft=flow_units_power_shaft,
            flow_units_pressure_ratio=flow_units_pressure_ratio,
            flow_units_disch_T=flow_units_disch_T,
            eff_units=eff_units,
            speed_units=speed_units,
            pressure_ratio_units=pressure_ratio_units,
            disch_T_units=disch_T_units,
            **curves_path_dict,
        )

    def save(self, file):
        """Save impeller to a toml file.

        Parameters
        ----------
        file : str or pathlib.Path
            Filename to which the data is saved.
        """

        with open(file, mode="w") as f:
            # add points to file
            dict_to_save = self._dict_to_save()
            toml.dump(dict_to_save, f)

    def _dict_to_save(self):
        dict_to_save = {
            f"Point{i}": point._dict_to_save() for i, point in enumerate(self.points)
        }
        return dict_to_save

    @classmethod
    def load(cls, file):
        """Load impeller from toml file.

        Parameters
        ----------
        file : str or pathlib.Path
            Filename to which the data is saved.

        Returns
        -------
        impeller : ccp.Impeller
            Impeller object.
        """
        parameters = toml.load(file)
        points = [
            Point(**Point._dict_from_load(kwargs)) for kwargs in parameters.values()
        ]

        return cls(points)

    def save_isis_txt(self, file, parameter):
        """Save curves to a csv with isis format.

        Parameters
        ----------
        file : str or pathlib.Path
            Filename to which the data is saved.
        parameter : str
            Parameter name. Can be "head", "eff", "power", "pressure_ratio" or "disch_T".

        Examples
        --------
        >>> impeller.save_isis_txt("head.txt", parameter="head")
        """
        # create dict with curves

        parameters_units = {
            "head": "kJ/kg",
            "eff": "percent",
        }
        curves_dict = {}
        for curve in self.curves:
            curves_dict[round(curve.speed.to("RPM").m)] = {
                "flow": getattr(curve, "flow_v").m,
                f"{parameter}": getattr(curve, parameter)
                .to(parameters_units[parameter])
                .m,
            }

        with open(file, mode="w", newline="") as f:
            for idx, (key, values) in enumerate(curves_dict.items(), start=1):
                f.write(f"curva{idx} {key}\n")
                for flow, param in zip(values["flow"], values[parameter]):
                    f.write(f"{flow:.5f} {param:.5f}\n")
                f.write("\n")  # empty line after each curve

    def save_hysys_csv(self, curve_dir):
        """Save curve to a csv with hysys format.

        curve_path: pathlib.Path
            Path for directory where the files will be saved.
        """
        curve_dir.mkdir(parents=True, exist_ok=True)
        surge = {"Speed (RPM)": [], "Volume Flow (m3/h)": []}
        stonewall = {"Speed (RPM)": [], "Volume Flow (m3/h)": []}

        for curve in self.curves:
            curve.save_hysys_csv(
                curve_dir / f'speed-{curve.speed.to("RPM").m:.0f}-RPM.csv'
            )
            surge["Speed (RPM)"].append(curve.speed.to("RPM").m)
            stonewall["Speed (RPM)"].append(curve.speed.to("RPM").m)
            surge["Volume Flow (m3/h)"].append(
                min(p.flow_v.to("m**3/h").m for p in curve)
            )
            stonewall["Volume Flow (m3/h)"].append(
                max(p.flow_v.to("m**3/h").m for p in curve)
            )

        for name, limit in zip(["surge", "stonewall"], [surge, stonewall]):
            with open(str(curve_dir / (name + ".csv")), mode="w") as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=list(limit.keys()))
                writer.writeheader()
                for speed, flow in zip(
                    limit["Speed (RPM)"], limit["Volume Flow (m3/h)"]
                ):
                    writer.writerow({"Speed (RPM)": speed, "Volume Flow (m3/h)": flow})


def find_closest_speeds(array, value):
    diff = array - value
    idx = np.abs(diff).argmin()

    if idx == 0:
        return [0, 1]
    elif idx == len(array) - 1:
        return [len(array) - 2, len(array) - 1]

    if diff[idx] > 0:
        idx = [idx - 1, idx]
    else:
        idx = [idx, idx + 1]

    return np.array(idx)


def get_interpolated_values(fac, flow_0, val_0, flow_1, val_1):
    x0 = [flow_0, val_0]
    if fac > 0.5:
        x0 = [flow_1, val_1]

    result = fsolve(system_to_interpolate, x0, args=(fac, flow_0, val_0, flow_1, val_1))

    flow_x = result[0]
    val_x = result[1]

    return flow_x, val_x


def system_to_interpolate(x, *args):
    fac, flow_0, val_0, flow_1, val_1 = args
    eq_1 = -x[1] + val_0 + (x[0] - flow_0) * (val_1 - val_0) / (flow_1 - flow_0)
    eq_2 = -fac * (flow_1 - flow_0) + x[0] - flow_0

    return [eq_1, eq_2]


def calculate_power_losses(power_losses_ref, speed_ref, speed):
    return power_losses_ref * (speed / speed_ref) ** 2.5


def impeller_example():
    test_dir = Path(__file__).parent / "tests"
    data_dir = test_dir / "data"

    suc = State(
        p=Q_(4.08, "bar"),
        T=Q_(33.6, "degC"),
        fluid={
            "METHANE": 58.976,
            "ETHANE": 3.099,
            "PROPANE": 0.6,
            "N-BUTANE": 0.08,
            "I-BUTANE": 0.05,
            "N-PENTANE": 0.01,
            "I-PENTANE": 0.01,
            "NITROGEN": 0.55,
            "HYDROGEN SULFIDE": 0.02,
            "CARBON DIOXIDE": 36.605,
        },
    )
    imp = Impeller.load_from_engauge_csv(
        suc=suc,
        curve_name="lp-sec1-caso-a",
        curve_path=data_dir,
        b=Q_(5.7, "mm"),
        D=Q_(550, "mm"),
        head_units="kJ/kg",
        flow_units="m³/h",
        number_of_points=7,
    )

    return imp


def converter(x):
    """Helper function used to parallelize conversion of points."""
    point, suc, find = x
    return Point.convert_from(point, suc=suc, find=find)


def create_points_parallel(x):
    """Helper function used to parallelize creation of points."""
    return Point(**x)


def calc_min_head_point(x, speed, imp, min_head):
    """Helper function to calculate min_head point."""
    try:
        p = imp.point(flow_v=x, speed=speed)
        head = p.head.m
    # if state does not converge, force head value just to keep the iterations going
    except ValueError:
        head = -x + min_head
    return head - min_head
